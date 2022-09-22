from openpyxl import load_workbook
from openpyxl.worksheet import worksheet
from openpyxl.workbook import workbook

class xlDicts:
    def __init__(self, wb: str = "", data_only: bool = True, auto_save: bool = False):
        self.file = wb
        self.wb: workbook = load_workbook(wb,  data_only=data_only)
        self.autoSave = auto_save
        self.data = dict()
        self.struct = dict()
        self.structuredData = dict()

    def __del__(self):
        self.wb.close()

    def __getSht(self, sht: str = "") -> worksheet:
        if len(sht) > 0:
            return self.wb[sht]
        else:
            return self.wb.active

    def load(self, sht: str = "", keyCol: int = 1, valCol = 2, startRow : int = 1, endRow: int = None, ignoreNullVal: bool = True, setNullValTo = 0, reversed:bool = False, asFormula:bool = False):
        """
                   load the data on the spreadsheet to dict
                   @param
                           sht             sht name, active sheet by default
                           keyCol          can be a single value           -> the normal Dicts instance
                           valCol          can be a single value           -> the normal Dicts instance
                                           or tuple with two elements      -> (keyColFrom, keyColTo)
                                           or list with more than one,     -> [keyLvl1, keyLvl2, ...]
                                           which specified the order of the columns
                           startRow        from which row to start
                           endRow          ends at which row
                           ignoreNullVal   whether to ignore the key if the corresponding value is None, default True
                                           if any one of the values is not None, the key will not be dropped
                           setNullValTo    set the None value to the given value
                           reversed        read from top to bottom. the duplicated value will be replaced by the one below
               """
        s = self.__getSht(sht)
        end = endRow or s.max_row

        if not type(valCol) is list:
            if type(valCol) is tuple:
                valCol = range(valCol[0], valCol[1] + 1)
            elif isinstance(valCol, int):
                valCol = [valCol]
            else:
                raise TypeError("the type of {0} is invalid!\n Integer, Tuple with two elements or List is required.")

        filterCol = [i - min(valCol) for i in valCol]
        k = list(list(s.iter_cols(keyCol, keyCol, startRow, end, True))[0])
        if asFormula:
            t = s.title
            v = [[i[j] for j in filterCol] for i in
                [
                [f"'{t}'!{s.cell(y, x).column_letter}{y}" for x in range(min(valCol), max(valCol) + 1)]
                    for y in range(startRow, end + 1)]
                ]
        else:
            v = [[i[j] for j in filterCol] for i in s.iter_rows(startRow, end, min(valCol), max(valCol), True)]

        if reversed:
            k.reverse()
            v.reverse()

        self.data = {k: v for k, v in dict(zip(k, v)).items() if not k is None}

        if not asFormula:
            if ignoreNullVal:
                self.data = {k: [setNullValTo if i is None else i for i in v] for k, v in self.data.items() if any([i is not None for i in v])}
            else:
                self.data = {k: [setNullValTo if i is None else i for i in v] for k, v in self.data.items()}

        if len(valCol) == 1:
            self.data = {k: v[0] for k, v in self.data.items()}

        return self

    def loadStruct(self, sht: str = "", keyCol: int = 1, valCol: int = 2):
        s = self.__getSht(sht)
        d = zip(list(list(s.iter_cols(keyCol, keyCol, None, None, True))[0]),
                list(list(s.iter_cols(valCol, valCol, None, None, True))[0]))
        acc = dict()
        tmp = list()
        lastK = None
        for k, v in d:
            if k is not None and v is None:
                if lastK is not None:
                    acc[lastK] = tmp
                lastK = k
                tmp = []
            elif k is None and v is not None:
                tmp.append(v)

        acc[lastK] = tmp
        self.struct = acc

        return self

    def feed(self, data=None):
        if not type(data) is dict:
            if data is None:
                data = self.data
            elif type(data) is xlDicts:
                data = data.data
            else:
                raise TypeError("the type of {0} is invalid!\n xlDicts or dict is required.")

        self.structuredData = {k: {k0: data[k0] for k0 in v if k0 in data} for k, v in self.struct.items()}
        return self

    def aggregate(self, func=lambda dic: sum(dic.values())):
        self.data = {k: func(v) for k, v in self.structuredData.items()}
        return self

    def fromDict(self, d: dict):
        self.data = d
        return self

    def unload(self, sht: str = "", keyCol: int = 1, startCol: int = 2, startRow: int = 1, endRow: int = None):
        if len(self.data) > 0:
            s: worksheet = self.__getSht(sht)
            end = endRow or s.max_row
            ks = list(self.data.keys())
            for r in range(startRow, end + 1):
                k = s.cell(column=keyCol, row=r).value
                if k in ks:
                    if not type(self.data[k]) is list:
                        tmp = [self.data[k]]
                    else:
                        tmp = self.data[k]

                    cnt = 0
                    for c in range(startCol, startCol + len(tmp)):
                        s.cell(column=c, row=r, value=tmp[cnt])
                        cnt = cnt + 1

            if self.autoSave:
                self.save()

        return self

    def dump(self, sht: str = "", topRow: int = 1, leftCol: int = 1):
        if len(self.data) > 0:
            s: worksheet = self.__getSht(sht)
            r = topRow

            for k, v in self.data.items():
                if not type(v) is list:
                    l = [v]
                else:
                    l = v

                cnt = 0
                s.cell(column=leftCol, row=r, value=k)
                for c in range(leftCol+1, leftCol + 1 + len(l)):
                    s.cell(column=c, row=r, value=l[cnt])
                    cnt = cnt + 1

                r = r + 1

            if self.autoSave:
                self.save()

        return self

    def dumpStructuredData(self, sht: str = "", topRow: int = 1, leftCol: int = 1):
        if len(self.structuredData) > 0:
            s: worksheet = self.__getSht(sht)

            r = topRow

            for k0, v0 in self.structuredData.items():
                s.cell(column=leftCol, row=r, value=k0)
                r = r + 1
                for k, v in v0.items():
                    s.cell(column=leftCol + 1, row=r, value=k)
                    cnt = 1
                    for r0 in v:
                        s.cell(column=leftCol + 1 + cnt, row=r, value=r0)
                        cnt = cnt + 1

                    r = r + 1

            if self.autoSave:
                self.save()

        return self

    def dumpStruct(self, sht: str = "", topRow: int = 1, leftCol: int = 1):
        if len(self.struct) > 0:
            s: worksheet = self.__getSht(sht)

            r = topRow

            for k, v in self.struct:
                s.cell(column=leftCol, row=r, value=k)
                r = r + 1
                for r0 in v:
                    s.cell(column=leftCol+1, row=r, value=r0)
                    r = r + 1

            if self.autoSave:
                self.save()

        return self

    def save(self):
        self.wb.save(self.file)
        return self






